import requests
import json
import pandas as pd
import numpy as np
import pprint
import time
from openpyxl import Workbook
from urllib.error import HTTPError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

_BASE_URL = "https://ubc.instructure.com/api/v1"

def make_request(url, token, method="GET", post_fields={}):
    request = Request(
        "{base_url}/{call_url}".format(base_url=_BASE_URL, call_url=url))
    request.add_header('Authorization', 'Bearer {token}'.format(token=token))
    request.method = method
    if post_fields:
        request.data = urlencode(post_fields).encode()
    try:
        response = urlopen(request)
    except HTTPError as e:
        return
    decoded_response = response.readline().decode("utf-8")
    response_body = json.loads(decoded_response, object_pairs_hook=dict)
    return response_body

def get_rubric_criterias(course_id, assignment_id, token):

    assignment = make_request("courses/{course_id}/assignments/{assignment_id}"
                .format(course_id=course_id, assignment_id= assignment_id), token, method = "GET")
    rubric = assignment['rubric']

    rubric_criteria = []

    for criteria in rubric:
        rubric_criteria.append(criteria['description'])

    return rubric_criteria



def get_students(student_ids,
                 user_request,
                 student_list,
                 log,
                 url,
                 course_id,
                 token):
    
    pagination = False
    counter = 0

    for student in student_list:
        student_ids[student['id']] = student['name']
        counter += 1
    
    while user_request.links['current']['url'] != user_request.links['last']['url']:
        user_request = requests.get(user_request.links['next']['url'],
                                    headers= {'Authorization': 'Bearer ' + token},
                                    timeout = 60)
        student_list = json.loads(user_request.text)

        for student in student_list:
            student_ids[student['id']] = student['name']
            counter += 1

    current_time = time.strftime("%B %d, %Y %I:%M:%S %p")
    log.write(current_time + ": " + str(counter) + " students gathered\n")    
    print('Students of the course have been extracted')
    return student_ids

def get_submissions(student_ids,
                    student_submissions,
                    comments_received,
                    submission_request,
                    submissions,
                    log,
                    url,
                    course_id,
                    token):

    feedback = {}
    counter = 0

    for sub in submissions:
        try:
            student_submissions[sub['id']] = student_ids[sub['user_id']]
            counter += 1
        except KeyError:
            current_time = time.strftime("%B %d, %Y %I:%M:%S %p")
            log.write(current_time + ": student not found\n") 
            continue
    
        for comment in sub['submission_comments']:
            comments_received[student_ids[sub['user_id']]][comment['author_name']] = {}
            comments_received[student_ids[sub['user_id']]][comment['author_name']]['Comments'] = comment['comment']

    while submission_request.links['current']['url'] != submission_request.links['last']['url']:
        submission_request = requests.get(submission_request.links['next']['url'],
                                  headers= {'Authorization': 'Bearer ' + token},
                                  timeout = 60)
        submissions = json.loads(submission_request.text)
        for sub in submissions:
            try:
                student_submissions[sub['id']] = student_ids[sub['user_id']]
                counter += 1
            except KeyError:
                current_time = time.strftime("%B %d, %Y %I:%M:%S %p")
                log.write(current_time + ": student not found\n") 
                continue
            
            for comment in sub['submission_comments']:
                comments_received[student_ids[sub['user_id']]][comment['author_name']] = {}
                comments_received[student_ids[sub['user_id']]][comment['author_name']]['Comments'] = comment['comment']

    current_time = time.strftime("%B %d, %Y %I:%M:%S %p")
    log.write(current_time + ": " + str(counter) + " submissions gathered\n")         
    print('Submissions for the assignment have been processed')
    return student_submissions

def get_comments(student_ids,
                 student_submissions,
                 comments_received,
                 rubric_criterias,
                 rubric,
                 rubric_json,
                 log,
                 url,
                 course_id,
                 token):
    peer_review = {}
    comment = {}
    score = {}
    total = {}
    counter = 0
    
    for assessment in rubric_json['assessments']:
        try:
            student_name = student_submissions[assessment['artifact_id']]
            commenter_name = student_ids[assessment['assessor_id']]
        except:
            continue
        
        try:
            comments_received[student_name][commenter_name]['Total Grade Received'] =  assessment['score']
        except KeyError:
            comments_received[student_name][commenter_name] = {}
            comments_received[student_name][commenter_name]['Comments'] = ''
            comments_received[student_name][commenter_name]['Total Grade Received'] =  assessment['score']
                    
        for i in range(len(assessment['data'])):
            comments_received[student_name][commenter_name][rubric_criterias[i] + ' Comment'] = assessment['data'][i]['comments']
            try:
                comments_received[student_name][commenter_name][rubric_criterias[i] + ' Score'] = assessment['data'][i]['points']
            except KeyError:
                comments_received[student_name][commenter_name][rubric_criterias[i] + ' Score'] = 0.0                                
        counter += 1
    current_time = time.strftime("%B %d, %Y %I:%M:%S %p")
    log.write(current_time + ": " + str(counter) + " assessments processed\n") 
    print("Submission Comments extracted")
    return comments_received

def calc_median(comments_received,
                log):
    grades = []
    for student in comments_received.keys():
        for commenter in comments_received[student].keys():
            try:
                grade = comments_received[student][commenter]['Total Grade Received']
                if grade != None:
                    grades.append(grade)
            except KeyError:
                current_time = time.strftime("%B %d, %Y %I:%M:%S %p")
                log.write(current_time + ": No grades found for " + student + "\n") 
                continue

        median = np.median(grades)
        if median > 0:
            comments_received[student]['Median of Grades'] = float(median)
        else:
            comments_received[student]['Median of Grades'] =  0
        grades = []
        
    current_time = time.strftime("%B %d, %Y %I:%M:%S %p")
    log.write(current_time + ": Median grades calculated\n") 


def calc_mean(comments_received,
                log):
    grades = []
    for student in comments_received.keys():
        for commenter in comments_received[student].keys():
            try:
                grade = comments_received[student][commenter]['Total Grade Received']
                if grade != None:
                    grades.append(grade)
            except KeyError:
                current_time = time.strftime("%B %d, %Y %I:%M:%S %p")
                log.write(current_time + ": No grades found for " + student + "\n") 
                continue

        mean = np.mean(grades)
        if mean > 0:
            comments_received[student]['Mean of Grades'] = float(mean)
        else:
            comments_received[student]['Mean of Grades'] =  0
        grades = []
        
    current_time = time.strftime("%B %d, %Y %I:%M:%S %p")
    log.write(current_time + ": Mean grades calculated\n")

def create_spreadsheet(comments_received,
                       column_letters, choice):
    s=""
    if(choice=='1'):
        s="Mean of Grades"
    else:
        s="Median of Grades"

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Student Names"
    ws["B1"] = s
    ws["C1"] = "Commenter Names"

    flag = 0
    cell_col = 2
    cell_row = 2
    for student in comments_received.keys():
        cell = column_letters[0] + str(cell_row)
        ws[cell] = student
        cell = column_letters[1] + str(cell_row)
        ws[cell] = comments_received[student][s]
        
        for commenter in comments_received[student].keys():
            if commenter ==  s:
                continue
            else:
                cell = column_letters[cell_col] + str(cell_row)
                ws[cell] = commenter
                cell_col += 1

            for header in comments_received[student][commenter].keys():
                if flag == 0:
                    cell = column_letters[cell_col] + "1"
                    ws[cell] = header
                cell = column_letters[cell_col] + str(cell_row)
                ws[cell] = comments_received[student][commenter][header]
                cell_col += 1
                
            flag = 1
            cell_row += 1
            cell_col = 2

    return wb
    
        
